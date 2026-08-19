from flask import Flask, render_template, request, redirect, send_file, jsonify, session, url_for
from datetime import datetime
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
app = Flask(__name__)
app.secret_key = "smart-finance-secret-key"
app.secret_key = "smartfinance123"

users = {}
current_user = {}

income_data = []
expense_data = []
investment_data = []
goal_data = []
def total_income():
    return round(sum(float(item["amount"]) for item in income_data), 2)
def generate_advanced_reports():

    income = round(float(total_income()), 2)

    expense = round(float(total_expense()), 2)

    savings = round(float(total_savings()), 2)

    expense_categories = {}

    for item in expense_data:

        category = item.get("category", "Other")

        amount = float(item.get("amount", 0))

        expense_categories[category] = (
            expense_categories.get(category, 0) + amount
        )

    expense_report = []

    for category, amount in expense_categories.items():

        percentage = 0

        if expense > 0:
            percentage = round((amount / expense) * 100, 2)

        expense_report.append({
            "category": category,
            "amount": round(amount, 2),
            "percentage": percentage
        })

    budget_report = []

    if income > 0:

        budget_categories = {
            "Needs": income * 0.50,
            "Wants": income * 0.30,
            "Savings": income * 0.20
        }

        needs_actual = 0
        wants_actual = 0

        for item in expense_data:

            category = item.get("category", "").lower()

            amount = float(item.get("amount", 0))

            if category in [
                "rent",
                "food",
                "bills",
                "utilities",
                "healthcare",
                "education"
            ]:

                needs_actual += amount

            else:

                wants_actual += amount

        actual_values = {
            "Needs": needs_actual,
            "Wants": wants_actual,
            "Savings": savings
        }

        for category, budget in budget_categories.items():

            actual = actual_values.get(category, 0)

            utilization = 0

            if budget > 0:
                utilization = round(
                    (actual / budget) * 100,
                    2
                )

            budget_report.append({
                "category": category,
                "budget": round(budget, 2),
                "actual": round(actual, 2),
                "utilization": utilization
            })

    investment_report = []

    investment_value = 0

    for item in investment_data:

        invested = float(item.get("invested", 0))

        current = float(item.get("current", 0))

        investment_value += current

        roi = 0

        if invested > 0:

            roi = round(
                ((current - invested) / invested) * 100,
                2
            )

        investment_report.append({
            "name": item.get("name", "Investment"),
            "invested": round(invested, 2),
            "current": round(current, 2),
            "roi": roi
        })

    goal_report = []

    for goal in goal_data:

        target = float(goal.get("target", 0))

        saved = float(goal.get("saved", 0))

        progress = 0

        if target > 0:

            progress = round(
                (saved / target) * 100,
                2
            )

        progress = min(progress, 100)

        goal_report.append({
            "name": goal.get("name", "Goal"),
            "target": round(target, 2),
            "saved": round(saved, 2),
            "progress": progress
        })

    return {
        "income": round(income, 2),
        "expense": round(expense, 2),
        "savings": round(savings, 2),
        "investment_value": round(investment_value, 2),
        "expense_report": expense_report,
        "budget_report": budget_report,
        "investment_report": investment_report,
        "goal_report": goal_report
    }
def total_expense():
    return round(sum(float(item["amount"]) for item in expense_data), 2)


def total_savings():
    return round(total_income() - total_expense(), 2)


def savings_rate():

    income = total_income()

    if income == 0:
        return 0

    return round((total_savings() / income) * 100, 2)
@app.route("/")
def home():

    if current_user == {}:
        return redirect("/login")

    return render_template(
        "home.html",
        user=current_user
    )
@app.route("/advanced_reports")
def advanced_reports():

    if current_user == {}:
        return redirect("/login")

    report = generate_advanced_reports()

    return render_template(
        "advanced_reports.html",
        user=current_user,
        **report
    )
@app.route("/export_report_pdf")
def export_report_pdf():

    if current_user == {}:
        return redirect("/login")

    report = generate_advanced_reports()

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "Smart Finance Insights",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "Advanced Financial Report",
            styles["Heading2"]
        )
    )

    elements.append(Spacer(1, 15))

    summary_data = [
        ["Financial Metric", "Amount"],
        ["Total Income", f"₹{report['income']:.2f}"],
        ["Total Expense", f"₹{report['expense']:.2f}"],
        ["Total Savings", f"₹{report['savings']:.2f}"],
        ["Investment Value", f"₹{report['investment_value']:.2f}"]
    ]

    summary_table = Table(summary_data)

    summary_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold")
        ])
    )

    elements.append(summary_table)

    elements.append(Spacer(1, 20))

    elements.append(
        Paragraph(
            "Expense Report",
            styles["Heading2"]
        )
    )

    expense_data = [
        ["Category", "Amount", "Percentage"]
    ]

    for item in report["expense_report"]:

        expense_data.append([
            item["category"],
            f"₹{item['amount']:.2f}",
            f"{item['percentage']}%"
        ])

    expense_table = Table(expense_data)

    expense_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER")
        ])
    )

    elements.append(expense_table)

    elements.append(Spacer(1, 20))

    elements.append(
        Paragraph(
            "Investment Performance",
            styles["Heading2"]
        )
    )

    investment_data_pdf = [
        ["Investment", "Invested", "Current", "ROI"]
    ]

    for item in report["investment_report"]:

        investment_data_pdf.append([
            item["name"],
            f"₹{item['invested']:.2f}",
            f"₹{item['current']:.2f}",
            f"{item['roi']}%"
        ])

    investment_table = Table(investment_data_pdf)

    investment_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER")
        ])
    )

    elements.append(investment_table)

    elements.append(Spacer(1, 20))

    elements.append(
        Paragraph(
            "Financial Goal Progress",
            styles["Heading2"]
        )
    )

    goal_data_pdf = [
        ["Goal", "Target", "Saved", "Progress"]
    ]

    for item in report["goal_report"]:

        goal_data_pdf.append([
            item["name"],
            f"₹{item['target']:.2f}",
            f"₹{item['saved']:.2f}",
            f"{item['progress']}%"
        ])

    goal_table = Table(goal_data_pdf)

    goal_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER")
        ])
    )

    elements.append(goal_table)

    document.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="smart_finance_report.pdf",
        mimetype="application/pdf"
    )
@app.route("/export_report_excel")
def export_report_excel():

    if current_user == {}:
        return redirect("/login")

    report = generate_advanced_reports()

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Financial Summary"

    sheet.append([
        "Financial Metric",
        "Amount"
    ])

    sheet.append([
        "Total Income",
        report["income"]
    ])

    sheet.append([
        "Total Expense",
        report["expense"]
    ])

    sheet.append([
        "Total Savings",
        report["savings"]
    ])

    sheet.append([
        "Investment Value",
        report["investment_value"]
    ])

    for cell in sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    expense_sheet = workbook.create_sheet(
        "Expense Report"
    )

    expense_sheet.append([
        "Category",
        "Amount",
        "Percentage"
    ])

    for item in report["expense_report"]:

        expense_sheet.append([
            item["category"],
            item["amount"],
            item["percentage"]
        ])

    investment_sheet = workbook.create_sheet(
        "Investment Report"
    )

    investment_sheet.append([
        "Investment",
        "Invested",
        "Current Value",
        "ROI"
    ])

    for item in report["investment_report"]:

        investment_sheet.append([
            item["name"],
            item["invested"],
            item["current"],
            item["roi"]
        ])

    goal_sheet = workbook.create_sheet(
        "Goal Report"
    )

    goal_sheet.append([
        "Goal",
        "Target",
        "Saved",
        "Progress"
    ])

    for item in report["goal_report"]:

        goal_sheet.append([
            item["name"],
            item["target"],
            item["saved"],
            item["progress"]
        ])

    for current_sheet in workbook.worksheets:

        for cell in current_sheet[1]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        for column in current_sheet.columns:

            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            current_sheet.column_dimensions[
                column_letter
            ].width = max_length + 3

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="smart_finance_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@app.route("/register", methods=["GET", "POST"])
def register():

    global users

    if request.method == "POST":

        username = request.form["username"]
        email = request.form["email"]
        password = request.form["password"]
        confirm = request.form["confirm_password"]

        if email in users:

            return render_template(
                "register.html",
                error="Email already registered."
            )

        if password != confirm:

            return render_template(
                "register.html",
                error="Passwords do not match."
            )

        users[email] = {

            "username": username,

            "email": email,

            "password": password

        }

        return redirect("/success")

    return render_template("register.html")
@app.route("/success")
def success():

    return render_template("success.html")
@app.route("/login", methods=["GET", "POST"])
def login():

    global current_user

    if request.method == "POST":

        email = request.form["email"]
        password = request.form["password"]

        if email not in users:

            return render_template(
                "login.html",
                error="Email not registered."
            )

        if users[email]["password"] != password:

            return render_template(
                "login.html",
                error="Incorrect password."
            )

        current_user = {

            "username": users[email]["username"],
            "email": email

        }

        return redirect("/dashboard")

    return render_template("login.html")
@app.route("/logout")
def logout():

    global current_user

    current_user = {}

    return redirect("/login")

@app.route("/dashboard")
def dashboard():

    if current_user == {}:
        return redirect("/login")

    income=total_income()

    expense=total_expense()

    savings=total_savings()

    invested=sum(i["invested"] for i in investment_data)

    score=financial_health_score()

    emergency=expense*6

    roi=0

    current=sum(i["current"] for i in investment_data)

    if invested>0:

        roi=round(((current-invested)/invested)*100,2)

    return render_template(

        "dashboard.html",

        user=current_user,

        total_income=income,

        total_expense=expense,

        total_savings=savings,

        savings_rate=savings_rate(),

        total_investment=invested,

        roi=roi,

        financial_score=score,

        emergency_fund=emergency,

        income_data=income_data,

        expense_data=expense_data,

        investment_data=investment_data,

        goal_data=goal_data

    )
@app.route("/income")
def income():

    if current_user == {}:
        return redirect("/login")

    return render_template(

        "income.html",

        income_data=income_data,

        total_income=total_income(),

        user=current_user

    )
@app.route("/add_income", methods=["POST"])
def add_income():

    source = request.form["source"]

    amount = float(request.form["amount"])

    date = request.form["date"]

    income_data.append({

        "source": source,

        "amount": amount,

        "date": date

    })

    return redirect("/income")
@app.route("/delete_income/<int:index>")
def delete_income(index):

    if index < len(income_data):

        income_data.pop(index)

    return redirect("/income")
@app.route("/expense")
def expense():

    if current_user == {}:
        return redirect("/login")

    return render_template(

        "expense.html",

        expense_data=expense_data,

        total_expense=total_expense(),

        user=current_user

    )
@app.route("/add_expense", methods=["POST"])
def add_expense():

    category = request.form["category"]

    amount = float(request.form["amount"])

    date = request.form["date"]

    expense_data.append({

        "category": category,

        "amount": amount,

        "date": date

    })

    return redirect("/expense")
@app.route("/delete_expense/<int:index>")
def delete_expense(index):

    if index < len(expense_data):

        expense_data.pop(index)

    return redirect("/expense")
def monthly_summary():

    income = total_income()

    expense = total_expense()

    savings = income - expense

    if income == 0:

        expense_ratio = 0

    else:

        expense_ratio = round((expense / income) * 100, 2)

    return {

        "income": income,

        "expense": expense,

        "savings": savings,

        "expense_ratio": expense_ratio

    }
@app.route("/reports")
def reports():

    if current_user == {}:
        return redirect("/login")

    summary = monthly_summary()

    return render_template(

        "reports.html",

        user=current_user,

        income_data=income_data,

        expense_data=expense_data,

        summary=summary

    )
@app.route("/charts")
def charts():

    if current_user == {}:
        return redirect("/login")

    categories = []
    amounts = []

    expense_summary = {}

    for item in expense_data:

        category = item["category"]

        if category not in expense_summary:

            expense_summary[category] = 0

        expense_summary[category] += item["amount"]

    for key, value in expense_summary.items():

        categories.append(key)

        amounts.append(value)

    return render_template(

        "charts.html",

        user=current_user,

        categories=categories,

        amounts=amounts,

        income=total_income(),

        expense=total_expense(),

        savings=total_savings()

    )
@app.route("/investment")
def investment():

    if current_user == {}:
        return redirect("/login")

    total_invested = sum(i["invested"] for i in investment_data)

    current_value = sum(i["current"] for i in investment_data)

    profit = current_value - total_invested

    roi = 0

    if total_invested > 0:

        roi = round((profit / total_invested) * 100, 2)

    return render_template(

        "investment.html",

        user=current_user,

        investment_data=investment_data,

        total_invested=total_invested,

        current_value=current_value,

        profit=profit,

        roi=roi

    )
@app.route("/add_investment", methods=["POST"])
def add_investment():

    asset = request.form["asset"]

    name = request.form["name"]

    invested = float(request.form["invested"])

    current = float(request.form["current"])

    investment_data.append({

        "asset": asset,

        "name": name,

        "invested": invested,

        "current": current

    })

    return redirect("/investment")
@app.route("/delete_investment/<int:index>")
def delete_investment(index):

    if index < len(investment_data):

        investment_data.pop(index)

    return redirect("/investment")
@app.route("/asset_allocation")
def asset_allocation():

    if current_user == {}:
        return redirect("/login")

    allocation = {}

    total = 0

    for item in investment_data:

        asset = item["asset"]

        invested = item["invested"]

        total += invested

        if asset not in allocation:

            allocation[asset] = 0

        allocation[asset] += invested

    allocation_data = []

    for asset, amount in allocation.items():

        percentage = 0

        if total > 0:

            percentage = round((amount / total) * 100, 2)

        allocation_data.append({

            "asset": asset,

            "amount": amount,

            "percentage": percentage

        })

    return render_template(

        "asset_allocation.html",

        user=current_user,

        allocation_data=allocation_data,

        total=total

    )
@app.route("/goals")
def goals():

    if current_user == {}:
        return redirect("/login")

    return render_template(

        "goals.html",

        user=current_user,

        goal_data=goal_data

    )
@app.route("/add_goal", methods=["POST"])
def add_goal():

    goal = request.form["goal"]

    target = float(request.form["target"])

    saved = float(request.form["saved"])

    date = request.form["date"]

    progress = 0

    if target > 0:

        progress = round((saved / target) * 100, 2)

    remaining = target - saved

    goal_data.append({

        "goal": goal,

        "target": target,

        "saved": saved,

        "remaining": remaining,

        "date": date,

        "progress": progress

    })

    return redirect("/goals")
@app.route("/delete_goal/<int:index>")
def delete_goal(index):

    if index < len(goal_data):

        goal_data.pop(index)

    return redirect("/goals")
@app.route("/analytics")
def analytics():

    if current_user == {}:
        return redirect("/login")

    income = total_income()

    expense = total_expense()

    savings = total_savings()

    invested = sum(i["invested"] for i in investment_data)

    current = sum(i["current"] for i in investment_data)

    roi = 0

    if invested > 0:
        roi = round(((current-invested)/invested)*100,2)

    return render_template(

        "analytics.html",

        user=current_user,

        income=income,

        expense=expense,

        savings=savings,

        savings_rate=savings_rate(),

        invested=invested,

        current=current,

        roi=roi,

        financial_score=financial_health_score(),

        goals=len(goal_data)

    )
@app.route("/profile")
def profile():

    if current_user == {}:
        return redirect("/login")

    income = total_income()

    expense = total_expense()

    savings = total_savings()

    invested = sum(i["invested"] for i in investment_data)

    return render_template(

        "profile.html",

        user=current_user,

        income=income,

        expense=expense,

        savings=savings,

        invested=invested,

        score=financial_health_score(),

        goals=len(goal_data)

    )
@app.route("/settings")
def settings():

    if current_user == {}:
        return redirect("/login")

    return render_template(

        "settings.html",

        user=current_user

    )
@app.route("/ai")
def ai():

    if current_user == {}:
        return redirect("/login")

    income = total_income()

    expense = total_expense()

    savings = total_savings()

    invested = sum(item["invested"] for item in investment_data)

    return render_template(

        "ai.html",

        user=current_user,

        income=income,

        expense=expense,

        savings=savings,

        invested=invested,

        goals=len(goal_data),

        score=financial_health_score()

    )
@app.route("/financial_health")
def financial_health():

    if current_user == {}:
        return redirect("/login")

    income = total_income()

    expense = total_expense()

    savings = total_savings()

    invested = sum(item["invested"] for item in investment_data)

    expense_ratio = 0
    savings_rate = 0
    investment_ratio = 0

    if income > 0:

        expense_ratio = round((expense/income)*100,2)

        savings_rate = round((savings/income)*100,2)

        investment_ratio = round((invested/income)*100,2)

    return render_template(

        "financial_health.html",

        user=current_user,

        income=income,

        expense=expense,

        savings=savings,

        invested=invested,

        expense_ratio=expense_ratio,

        savings_rate=savings_rate,

        investment_ratio=investment_ratio,

        score=financial_health_score()

    )
def financial_health_score():

    income = total_income()
    expense = total_expense()
    savings = total_savings()

    invested = sum(
        float(item.get("invested", 0) or 0)
        for item in investment_data
    )

    score = 0

    if income <= 0:
        return 0

    savings_rate = (savings / income) * 100
    expense_rate = (expense / income) * 100
    investment_ratio = (invested / income) * 100
    if savings_rate >= 40:
        score += 30

    elif savings_rate >= 30:
        score += 25

    elif savings_rate >= 20:
        score += 20

    elif savings_rate >= 10:
        score += 10

    else:
        score += 0
    if expense_rate <= 50:
        score += 30

    elif expense_rate <= 60:
        score += 25

    elif expense_rate <= 70:
        score += 20

    elif expense_rate <= 80:
        score += 10

    else:
        score += 0
    if investment_ratio >= 40:
        score += 20

    elif investment_ratio >= 30:
        score += 15

    elif investment_ratio >= 20:
        score += 10

    elif investment_ratio >= 10:
        score += 5

    else:
        score += 0
    if goal_data:

        total_goal_progress = 0
        valid_goals = 0

        for goal in goal_data:

            target = float(
                goal.get("target", 0) or 0
            )

            saved = float(
                goal.get("saved", 0) or 0
            )

            if target > 0:

                progress = (saved / target) * 100

                progress = min(progress, 100)

                total_goal_progress += progress

                valid_goals += 1

        if valid_goals > 0:

            average_goal_progress = (
                total_goal_progress / valid_goals
            )

            if average_goal_progress >= 80:
                score += 20

            elif average_goal_progress >= 60:
                score += 15

            elif average_goal_progress >= 40:
                score += 10

            elif average_goal_progress >= 20:
                score += 5

    return min(round(score), 100)

@app.route("/spending_analysis")
def spending_analysis():

    if current_user == {}:
        return redirect("/login")

    category_totals = {}

    for item in expense_data:

        category = item["category"]

        amount = float(item["amount"])

        category_totals[category] = category_totals.get(category, 0) + amount

    total = sum(category_totals.values())

    analysis = []

    highest_category = "None"

    highest_amount = 0

    for category, amount in category_totals.items():

        percentage = round((amount / total) * 100, 2) if total > 0 else 0

        analysis.append({

            "category": category,

            "amount": amount,

            "percentage": percentage

        })

        if amount > highest_amount:

            highest_amount = amount

            highest_category = category

    average_expense = round(total / len(category_totals), 2) if category_totals else 0

    return render_template(

        "spending_analysis.html",

        user=current_user,

        total_expense=total,

        categories=list(category_totals.keys()),

        analysis=analysis,

        highest_category=highest_category,

        average_expense=average_expense

    )
@app.route("/budget_recommendation")
def budget_recommendation():

    if current_user == {}:
        return redirect("/login")

    income = total_income()

    expense = total_expense()

    needs_budget = round(income * 0.50, 2)

    wants_budget = round(income * 0.30, 2)

    savings_budget = round(income * 0.20, 2)

    investment_budget = savings_budget

    return render_template(

        "budget_recommendation.html",

        user=current_user,

        income=income,

        expense=expense,

        needs_budget=needs_budget,

        wants_budget=wants_budget,

        savings_budget=savings_budget,

        investment_budget=investment_budget

    )
def jarvis_response(message):

    message = message.lower().strip()

    income = total_income()

    expense = total_expense()

    savings = total_savings()

    invested = sum(
        float(item.get("invested", 0) or 0)
        for item in investment_data
    )

    current = sum(
        float(item.get("current", 0) or 0)
        for item in investment_data
    )

    if income > 0:

        savings_rate_value = round(
            (savings / income) * 100,
            2
        )

        expense_rate_value = round(
            (expense / income) * 100,
            2
        )

    else:

        savings_rate_value = 0

        expense_rate_value = 0

    if invested > 0:

        roi = round(
            ((current - invested) / invested) * 100,
            2
        )

    else:

        roi = 0

    if (
        "income" in message
        or "salary" in message
        or "earn" in message
    ):

        return (
            f"Your total recorded income is "
            f"₹{income:,.2f}."
        )

    if (
        "expense" in message
        or "spend" in message
        or "spent" in message
    ):

        return (
            f"Your total recorded expenses are "
            f"₹{expense:,.2f}. "
            f"Your expense ratio is "
            f"{expense_rate_value}%."
        )

    if (
        "saving" in message
        or "save" in message
    ):

        return (
            f"Your current savings are "
            f"₹{savings:,.2f}, "
            f"which is {savings_rate_value}% "
            f"of your income."
        )

    if (
        "investment" in message
        or "investments" in message
        or "portfolio" in message
    ):

        return (
            f"Your total invested amount is "
            f"₹{invested:,.2f}. "
            f"Your current investment value is "
            f"₹{current:,.2f}. "
            f"Your calculated ROI is {roi}%."
        )

    if (
        "goal" in message
        or "goals" in message
        or "target" in message
    ):

        if not goal_data:

            return (
                "You currently have no financial goals. "
                "Create a goal to start tracking your progress."
            )

        responses = []

        for goal in goal_data:

            name = goal.get(
                "name",
                "Financial Goal"
            )

            target = float(
                goal.get("target", 0)
            )

            saved = float(
                goal.get("saved", 0)
            )

            progress = 0

            if target > 0:

                progress = min(
                    round(
                        (saved / target) * 100,
                        2
                    ),
                    100
                )

            responses.append(
                f"{name}: {progress}% completed"
            )

        return "Your goal progress: " + "; ".join(
            responses
        )

    if (
        "budget" in message
        or "50 30 20" in message
        or "50-30-20" in message
    ):

        needs = income * 0.50

        wants = income * 0.30

        recommended_savings = income * 0.20

        return (
            f"Based on the 50-30-20 rule, "
            f"your recommended budget is: "
            f"Needs ₹{needs:,.2f}, "
            f"Wants ₹{wants:,.2f}, "
            f"Savings ₹{recommended_savings:,.2f}."
        )

    if (
        "health" in message
        or "score" in message
        or "financial health" in message
    ):

        score = financial_health_score()

        if score >= 80:

            level = "Excellent"

        elif score >= 60:

            level = "Good"

        elif score >= 40:

            level = "Average"

        elif score >= 20:

            level = "Poor"

        else:

            level = "Critical"

        return (
            f"Your financial health score is "
            f"{score}/100. "
            f"Your current financial health level is "
            f"{level}."
        )

    if (
        "advice" in message
        or "recommend" in message
        or "suggestion" in message
    ):

        if savings_rate_value < 20:

            return (
                "Your savings rate is below 20%. "
                "Try reducing unnecessary expenses "
                "and increase your monthly savings."
            )

        if expense_rate_value > 80:

            return (
                "Your expenses are using more than "
                "80% of your income. "
                "Consider reducing discretionary spending."
            )

        if invested == 0:

            return (
                "You currently have no recorded investments. "
                "Consider learning about suitable investment "
                "options based on your financial goals."
            )

        return (
            "Your finances are currently in a reasonable "
            "position. Continue monitoring expenses, "
            "maintaining savings and reviewing investments."
        )

    if (
        "hello" in message
        or "hi" in message
        or "hey" in message
    ):

        return (
            "Hello! I'm JARVIS. "
            "I can help you understand your income, "
            "expenses, savings, investments, goals, "
            "budget and financial health."
        )

    return (
        "I can help with income, expenses, savings, "
        "investments, goals, budget and financial health. "
        "Please ask me a specific financial question."
    )
@app.route("/jarvis")
def jarvis():

    if current_user == {}:
        return redirect("/login")

    history = session.get("jarvis_history", [])

    return render_template(
        "jarvis.html",
        user=current_user,
        history=history
    )
@app.route("/jarvis_chat", methods=["POST"])
def jarvis_chat():

    if current_user == {}:

        return jsonify({
            "reply": "Please login first."
        }), 401

    data = request.get_json()

    if not data:

        return jsonify({
            "reply": "Please enter a message."
        }), 400

    message = str(
        data.get("message", "")
    ).strip()
    if len(message) < 2:

      return jsonify({
        "reply": "Please enter a meaningful question."
    }), 400

    if not message:

        return jsonify({
            "reply": "Please enter a financial question."
        }), 400

    if len(message) > 500:

        return jsonify({
            "reply": "Please keep your question below 500 characters."
        }), 400

    reply = jarvis_response(message)

    history = session.get(
        "jarvis_history",
        []
    )

    history.append({
        "user": message,
        "jarvis": reply
    })

    session["jarvis_history"] = history

    session.modified = True

    return jsonify({
        "reply": reply
    })
@app.route("/clear_jarvis", methods=["POST"])
def clear_jarvis():

    if current_user == {}:

        return jsonify({
            "message": "Please login first."
        }), 401

    session.pop(
        "jarvis_history",
        None
    )

    return jsonify({
        "message": "Conversation cleared."
    })
@app.route("/jarvis_summary")
def jarvis_summary():

    if current_user == {}:
        return jsonify({
            "error": "Please login first."
        }), 401

    income = total_income()

    expense = total_expense()

    savings = total_savings()

    invested = sum(
        float(item.get("invested", 0))
        for item in investment_data
    )

    current = sum(
        float(item.get("current", 0))
        for item in investment_data
    )

    roi = 0

    if invested > 0:

        roi = round(
            ((current - invested) / invested) * 100,
            2
        )

    return jsonify({

        "income": round(income, 2),

        "expense": round(expense, 2),

        "savings": round(savings, 2),

        "investment": round(current, 2),

        "roi": roi,

        "financial_health": financial_health_score(),

        "goals": len(goal_data)

    })
@app.route("/alerts")
def alerts():

    if current_user == {}:
        return redirect("/login")

    income = total_income()
    expense = total_expense()
    savings = total_savings()

    invested = sum(i["invested"] for i in investment_data)
    current = sum(i["current"] for i in investment_data)

    roi = 0
    expense_ratio = 0
    savings_rate = 0

    if invested > 0:
        roi = round(((current-invested)/invested)*100,2)

    if income > 0:
        expense_ratio = round((expense/income)*100,2)
        savings_rate = round((savings/income)*100,2)

    alerts = []

    if expense > income:
        alerts.append({
            "type":"warning",
            "message":"⚠ Your expenses are higher than your income."
        })

    if savings_rate < 20:
        alerts.append({
            "type":"info",
            "message":"💰 Try to save at least 20% of your income."
        })

    if invested == 0:
        alerts.append({
            "type":"warning",
            "message":"📈 Start investing to grow your wealth."
        })

    if len(goal_data) == 0:
        alerts.append({
            "type":"info",
            "message":"🎯 Create your first financial goal."
        })

    if financial_health_score() >= 80:
        alerts.append({
            "type":"success",
            "message":"🏆 Excellent financial health. Keep it up!"
        })

    success_count = len([a for a in alerts if a["type"]=="success"])
    info_count = len([a for a in alerts if a["type"]=="info"])
    warning_count = len([a for a in alerts if a["type"]=="warning"])

    return render_template(

        "alerts.html",

        user=current_user,

        alerts=alerts,

        roi=roi,

        savings_rate=savings_rate,

        expense_ratio=expense_ratio,

        success_count=success_count,

        info_count=info_count,

        warning_count=warning_count

    )
@app.errorhandler(404)
def page_not_found(error):

    if current_user == {}:
        return redirect("/login")

    return render_template(
        "error.html",
        user=current_user,
        error_code=404,
        error_message="The requested page was not found."
    ), 404


@app.errorhandler(500)
def internal_server_error(error):

    if current_user == {}:
        return redirect("/login")

    return render_template(
        "error.html",
        user=current_user,
        error_code=500,
        error_message="Something went wrong while processing your request."
    ), 500
if __name__=="__main__":

    app.run(debug=True)